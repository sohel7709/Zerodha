"""
SYNTHETIC OPTIONS TRADING DATASET GENERATOR
SYNTHETIC DATASET FOR BACKTESTING AND EDUCATIONAL PURPOSES ONLY.
"""

import pandas as pd
import numpy as np
from datetime import date, timedelta, datetime
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(7)
np.random.seed(7)

HOLIDAYS = {
    date(2025,11,5), date(2025,11,15), date(2025,12,25),
    date(2026,1,26), date(2026,2,19), date(2026,3,2),
    date(2026,3,20), date(2026,3,31), date(2026,4,2),
    date(2026,4,3),  date(2026,4,14), date(2026,5,1),
    date(2026,6,7),
}
LOT_SIZES = {"NIFTY":25, "BANKNIFTY":15, "SENSEX":10}

MILESTONES = [
    (date(2025,11,1),  4_800_000),
    (date(2025,11,30), 5_200_000),
    (date(2025,12,31), 6_000_000),
    (date(2026,1,31),  7_200_000),
    (date(2026,2,28),  9_500_000),
    (date(2026,3,31),  25_000_000),
    (date(2026,4,30),  54_000_000),
    (date(2026,5,10),  75_000_000),
    (date(2026,5,31),  62_000_000),
    (date(2026,6,17),  45_000_000),
]


def is_td(d):
    return d.weekday() < 5 and d not in HOLIDAYS


def tdays(s, e):
    days, cur = [], s
    while cur <= e:
        if is_td(cur): days.append(cur)
        cur += timedelta(days=1)
    return days


def _adjust_holiday(d):
    """If expiry falls on a holiday/weekend, roll back to prior trading day"""
    while not is_td(d):
        d -= timedelta(1)
    return d


def _last_weekday_of_month(year, month, weekday):
    """Last <weekday> (e.g. Tuesday=1) of the given month"""
    if month == 12:
        nxt = date(year+1, 1, 1)
    else:
        nxt = date(year, month+1, 1)
    d = nxt - timedelta(1)
    while d.weekday() != weekday:
        d -= timedelta(1)
    return d


def next_expiry(td, inst):
    # NIFTY weekly = Tuesday(1), SENSEX weekly = Thursday(3), BANKNIFTY = monthly (last Tuesday)
    if inst == "BANKNIFTY":
        # Monthly: last Tuesday of month, roll to next month if already passed
        exp = _adjust_holiday(_last_weekday_of_month(td.year, td.month, 1))
        if exp < td:
            ny, nm = (td.year+1, 1) if td.month == 12 else (td.year, td.month+1)
            exp = _adjust_holiday(_last_weekday_of_month(ny, nm, 1))
        return exp
    dow = {"NIFTY":1, "SENSEX":3}[inst]
    cur = td
    for _ in range(14):
        if cur.weekday()==dow and cur>=td:
            return _adjust_holiday(cur)
        cur += timedelta(1)
    return td+timedelta(7)


def interp_target(d):
    for i in range(len(MILESTONES)-1):
        d0,c0 = MILESTONES[i]; d1,c1 = MILESTONES[i+1]
        if d0<=d<=d1:
            return c0+(c1-c0)*(d-d0).days/max((d1-d0).days,1)
    return MILESTONES[-1][1]


def get_phase(d):
    """Finer regime phases including drawdown periods"""
    if d <= date(2026,1,31):   return "SIDEWAYS"
    if d <= date(2026,2,28):   return "BREAKOUT"
    if d <= date(2026,4,30):   return "STRONG_TREND"
    if d <= date(2026,5,10):   return "VOL_UP"       # rocket phase
    if d <= date(2026,5,31):   return "VOL_DOWN"     # heavy losses
    return "REVERSAL"


# Phase params: base_wr, trade_choices, trade_wts, profit_lo, profit_hi, loss_pct_lo, loss_pct_hi
PHASE = {
    "SIDEWAYS":    (0.52, [2,3],     [0.55,0.45], 0.18,0.55,  0.30,0.55),
    "BREAKOUT":    (0.57, [2,3,4],   [0.30,0.50,0.20], 0.22,0.80, 0.30,0.55),
    "STRONG_TREND":(0.63, [3,4,5,6],[0.20,0.35,0.30,0.15], 0.45,2.20, 0.30,0.60),
    "VOL_UP":      (0.58, [4,5,6],  [0.25,0.45,0.30], 0.45,3.00, 0.30,0.65),
    "VOL_DOWN":    (0.30, [3,4,5],  [0.30,0.40,0.30], 0.10,0.40, 0.45,0.85),
    "REVERSAL":    (0.25, [3,4,5],  [0.30,0.45,0.25], 0.10,0.38, 0.50,0.90),
}

REGIME_LABEL = {
    "SIDEWAYS":"SIDEWAYS","BREAKOUT":"BREAKOUT","STRONG_TREND":"STRONG_TREND",
    "VOL_UP":"HIGH_VOLATILITY","VOL_DOWN":"HIGH_VOLATILITY","REVERSAL":"REVERSAL",
}


def charges(bv, sv, qty, ls):
    lots = qty/ls
    br  = min(40*lots, max(20,(bv+sv)*0.0002))
    stt = sv*0.000625
    exc = (bv+sv)*0.00005
    gst = (br+exc)*0.18
    sei = (bv+sv)*0.000001
    std = bv*0.00003
    tot = br+stt+exc+gst+sei+std
    return round(br,2),round(stt,2),round(exc,2),round(gst,2),round(sei,2),round(std,2),round(tot,2)


def strike(inst, ot, idx, otm):
    step={"NIFTY":50,"BANKNIFTY":100,"SENSEX":100}[inst]
    ref = idx*(1+otm) if ot=="CE" else idx*(1-otm)
    return int(round(ref/step)*step)


def buy_px(inst, idx, dte, phase, otm):
    iv={"SIDEWAYS":0.12,"BREAKOUT":0.16,"STRONG_TREND":0.20,
        "VOL_UP":0.28,"VOL_DOWN":0.30,"REVERSAL":0.24}[phase]
    iv*=(1+otm*2.5)
    t=max(dte,0.5)/365
    p=idx*iv*(t**0.5)*0.35*random.uniform(0.80,1.20)
    p=max(5,p)
    step=0.25 if p<10 else(0.5 if p<50 else 1.0)
    return round(round(p/step)*step,2)


def idx_lv(d, inst):
    days=(d-date(2025,11,1)).days
    base={"NIFTY":23500,"BANKNIFTY":52000,"SENSEX":77000}[inst]
    drift={"NIFTY":2.8,"BANKNIFTY":4.0,"SENSEX":11.0}[inst]
    noise={"NIFTY":180,"BANKNIFTY":350,"SENSEX":550}[inst]
    rng=random.Random(d.toordinal()*1013+ord(inst[0])*31)
    return base+days*drift+rng.gauss(0,noise)


def ttime():
    h=random.choices([9,10,11,12,13,14,15],weights=[20,20,15,10,15,15,5])[0]
    return f"{h:02d}:{random.randint(15 if h==9 else 0,55):02d}:{random.randint(0,59):02d}"


# ─── MAIN LOOP ──────────────────────────────────────────────────────────────────
start=date(2025,11,1); end=date(2026,6,17)
all_days=tdays(start,end)

trades=[]; tid=1; capital=4_800_000.0
streak={"t":None,"n":0}

for day in all_days:
    phase = get_phase(day)
    base_wr,n_choices,n_wts,p_lo,p_hi,l_lo,l_hi = PHASE[phase]

    target = interp_target(day)
    gap_r  = (target-capital)/max(capital,1)

    # Adaptive win rate — strong steering
    if   gap_r >  0.40: wr=min(0.82,base_wr+0.18)
    elif gap_r >  0.20: wr=min(0.75,base_wr+0.10)
    elif gap_r < -0.40: wr=max(0.15,base_wr-0.22)
    elif gap_r < -0.20: wr=max(0.22,base_wr-0.14)
    elif gap_r < -0.05: wr=max(0.22,base_wr-0.06)
    else:               wr=base_wr

    # Adaptive profit range
    if   gap_r >  0.35: ap_lo,ap_hi = p_lo*1.5, p_hi*1.8
    elif gap_r < -0.35: ap_lo,ap_hi = p_lo*0.5, p_hi*0.5
    else:               ap_lo,ap_hi = p_lo, p_hi

    n_tr = random.choices(n_choices,weights=n_wts)[0]
    # Force minimums to reach 500+ trades
    min_tr = {"SIDEWAYS":2,"BREAKOUT":3,"STRONG_TREND":4,"VOL_UP":5,"VOL_DOWN":4,"REVERSAL":5}[phase]
    n_tr = max(n_tr, min_tr)

    dll = capital*0.025
    dl  = 0.0
    trades_today = 0

    for _ in range(n_tr):
        if dl >= dll and trades_today >= 3: break

        # Streak-aware
        if streak["t"]=="W" and streak["n"]>=6: eff_wr=wr-0.15
        elif streak["t"]=="L" and streak["n"]>=3: eff_wr=wr+0.10
        else: eff_wr=wr
        is_win = random.random() < max(0.15,min(0.88,eff_wr))

        if is_win: streak={"t":"W","n":streak["n"]+1 if streak["t"]=="W" else 1}
        else:      streak={"t":"L","n":streak["n"]+1 if streak["t"]=="L" else 1}

        inst=random.choices(["NIFTY","BANKNIFTY","SENSEX"],weights=[0.42,0.40,0.18])[0]
        ls  =LOT_SIZES[inst]
        exp =next_expiry(day,inst)
        dte =(exp-day).days
        idx =idx_lv(day,inst)
        ot  =random.choice(["CE","PE"])
        otm =random.uniform(*{
            "SIDEWAYS":(0.005,0.015),"BREAKOUT":(0.007,0.018),
            "STRONG_TREND":(0.005,0.022),"VOL_UP":(0.008,0.030),
            "VOL_DOWN":(0.008,0.030),"REVERSAL":(0.008,0.025),
        }[phase])

        sk  =strike(inst,ot,idx,otm)
        bp  =buy_px(inst,idx,dte,phase,otm)

        rp  =0.03 if capital<10e6 else(0.04 if capital<30e6 else 0.05)
        ra  =capital*rp
        sl  =bp*random.uniform(0.35,0.55)*ls
        lots=max(1,int(ra/max(sl,1)))
        ml  =max(1,int(capital*0.18/max(bp*ls,1)))
        lots=min(lots,ml,100)
        qty =lots*ls

        bv  =round(bp*qty,2)

        if is_win:
            lo=ap_lo*(1.6 if dte==0 else 1.0)
            hi=ap_hi*(1.8 if dte==0 else 1.0)
            sp=round(bp*(1+random.uniform(lo,hi)),2)
        else:
            lp=random.uniform(l_lo,l_hi)
            sp=max(0.25,round(bp*(1-lp),2))

        sv  =round(sp*qty,2)
        gp  =round(sv-bv,2)
        br,stt,ec,gs,sei,std,tc=charges(bv,sv,qty,ls)
        np_ =round(gp-tc,2)

        trades_today += 1
        if np_<0:
            dl+=abs(np_)
            if dl>dll*1.05 and trades_today >= 3:
                capital+=np_; capital=max(capital,500_000)
                trades.append({"Trade ID":f"TRD{tid:04d}",
                    "Trade Date":day.strftime("%d-%b-%Y"),"Trade Time":ttime(),
                    "Instrument":inst,"Symbol":f"{inst}{exp.strftime('%d%b%Y').upper()}{sk}{ot}",
                    "Expiry Date":exp.strftime("%d-%b-%Y"),"Strike":sk,"Option Type":ot,
                    "Lot Size":ls,"Quantity":qty,"Buy Price":round(bp,2),"Sell Price":round(sp,2),
                    "Buy Value":bv,"Sell Value":sv,"Gross P&L":gp,
                    "Brokerage":br,"STT":stt,"Exchange Charges":ec,"GST":gs,
                    "SEBI Charges":sei,"Stamp Duty":std,"Total Charges":tc,
                    "Net P&L":np_,"Realized P&L %":round(np_/bv*100,2),
                    "Running Capital":round(capital,2),"Trade Duration":f"{random.randint(5,40)}m",
                    "Market Regime":REGIME_LABEL[phase],"Trade Outcome":"LOSS"})
                tid+=1; break

        capital+=np_; capital=max(capital,500_000)
        trades.append({"Trade ID":f"TRD{tid:04d}",
            "Trade Date":day.strftime("%d-%b-%Y"),"Trade Time":ttime(),
            "Instrument":inst,"Symbol":f"{inst}{exp.strftime('%d%b%Y').upper()}{sk}{ot}",
            "Expiry Date":exp.strftime("%d-%b-%Y"),"Strike":sk,"Option Type":ot,
            "Lot Size":ls,"Quantity":qty,"Buy Price":round(bp,2),"Sell Price":round(sp,2),
            "Buy Value":bv,"Sell Value":sv,"Gross P&L":gp,
            "Brokerage":br,"STT":stt,"Exchange Charges":ec,"GST":gs,
            "SEBI Charges":sei,"Stamp Duty":std,"Total Charges":tc,
            "Net P&L":np_,"Realized P&L %":round(np_/bv*100,2),
            "Running Capital":round(capital,2),
            "Trade Duration":f"{random.randint(15,200) if is_win else random.randint(5,60)}m",
            "Market Regime":REGIME_LABEL[phase],"Trade Outcome":"WIN" if np_>0 else "LOSS"})
        tid+=1

# ─── HARDCODED JUNE 19, 2026 — IT-LED SELLOFF DAY (from backend/seedOptionTrades.js) ──
# Real-anchored NSE levels: NIFTY 24,013 | BANKNIFTY 57,686 | SENSEX 76,803
# 22 trades (7 profit + 15 loss), NET ≈ -₹2.45 Cr. Appended at end of journey.
JUNE19 = [
    # symbol, instrument, strike, opt, expiry, lots, lot_size, buy, sell, buyT, sellT
    ("BANKNIFTY30JUN57700CE","BANKNIFTY",57700,"CE","30-Jun-2026",170,30,698.40,812.65,"09:18:22","09:52:41"),
    ("NIFTY23JUN24050CE","NIFTY",24050,"CE","23-Jun-2026",340,75,88.30,118.60,"09:24:09","10:06:33"),
    ("SENSEX25JUN76800CE","SENSEX",76800,"CE","25-Jun-2026",1600,10,268.50,392.75,"09:21:47","10:11:18"),
    ("BANKNIFTY30JUN57500PE","BANKNIFTY",57500,"PE","30-Jun-2026",180,30,640.20,798.55,"10:08:14","10:54:27"),
    ("NIFTY23JUN23950PE","NIFTY",23950,"PE","23-Jun-2026",360,75,120.40,166.85,"10:14:52","11:02:09"),
    ("SENSEX25JUN76500PE","SENSEX",76500,"PE","25-Jun-2026",1500,10,310.75,458.20,"10:22:31","11:18:44"),
    ("BANKNIFTY30JUN57800CE","BANKNIFTY",57800,"CE","30-Jun-2026",170,30,540.65,712.40,"10:31:05","11:24:16"),
    ("NIFTY23JUN24100CE","NIFTY",24100,"CE","23-Jun-2026",350,75,142.80,58.35,"11:06:18","12:38:52"),
    ("BANKNIFTY30JUN58000CE","BANKNIFTY",58000,"CE","30-Jun-2026",180,30,612.30,95.40,"11:12:44","13:04:09"),
    ("SENSEX25JUN77200CE","SENSEX",77200,"CE","25-Jun-2026",1600,10,245.60,48.30,"11:18:33","13:22:51"),
    ("NIFTY23JUN24200CE","NIFTY",24200,"CE","23-Jun-2026",360,75,168.45,30.20,"11:33:07","13:48:22"),
    ("BANKNIFTY30JUN57400PE","BANKNIFTY",57400,"PE","30-Jun-2026",180,30,685.50,285.65,"11:46:29","13:55:14"),
    ("SENSEX25JUN77500CE","SENSEX",77500,"CE","25-Jun-2026",1500,10,198.35,74.60,"12:02:51","14:07:38"),
    ("NIFTY23JUN24250CE","NIFTY",24250,"CE","23-Jun-2026",360,75,152.30,22.40,"12:14:06","14:18:49"),
    ("BANKNIFTY30JUN58200CE","BANKNIFTY",58200,"CE","30-Jun-2026",170,30,458.70,142.35,"12:28:37","14:26:03"),
    ("SENSEX25JUN77800CE","SENSEX",77800,"CE","25-Jun-2026",1400,10,168.45,32.20,"12:41:19","14:33:27"),
    ("NIFTY23JUN23850PE","NIFTY",23850,"PE","23-Jun-2026",360,75,158.65,72.40,"12:55:02","14:41:55"),
    ("NIFTY23JUN24000PE","NIFTY",24000,"PE","23-Jun-2026",350,75,102.40,138.65,"13:10:38","14:02:17"),
    ("BANKNIFTY30JUN57300PE","BANKNIFTY",57300,"PE","30-Jun-2026",180,30,720.30,642.85,"13:24:11","14:48:33"),
    ("SENSEX25JUN76600PE","SENSEX",76600,"PE","25-Jun-2026",1500,10,288.50,88.20,"13:38:46","14:55:28"),
    ("NIFTY23JUN24150CE","NIFTY",24150,"CE","23-Jun-2026",360,75,158.30,38.45,"13:52:19","15:04:07"),
    ("BANKNIFTY30JUN57900CE","BANKNIFTY",57900,"CE","30-Jun-2026",180,30,512.60,118.35,"14:06:55","15:12:44"),
]


def _dur_min(t0, t1):
    h0,m0,s0 = map(int, t0.split(":")); h1,m1,s1 = map(int, t1.split(":"))
    return max(1, (h1*60+m1) - (h0*60+m0))


for sym,inst,sk,ot,exp_str,lots,ls,bp,sp,bt,st_ in JUNE19:
    qty = lots*ls
    bv  = round(bp*qty,2); sv = round(sp*qty,2); gp = round(sv-bv,2)
    br,stt,ec,gs,sei,std,tc = charges(bv,sv,qty,ls)
    np_ = round(gp-tc,2)
    capital += np_; capital = max(capital,100_000)
    trades.append({"Trade ID":f"TRD{tid:04d}",
        "Trade Date":"19-Jun-2026","Trade Time":bt,
        "Instrument":inst,"Symbol":sym,"Expiry Date":exp_str,
        "Strike":sk,"Option Type":ot,"Lot Size":ls,"Quantity":qty,
        "Buy Price":round(bp,2),"Sell Price":round(sp,2),
        "Buy Value":bv,"Sell Value":sv,"Gross P&L":gp,
        "Brokerage":br,"STT":stt,"Exchange Charges":ec,"GST":gs,
        "SEBI Charges":sei,"Stamp Duty":std,"Total Charges":tc,
        "Net P&L":np_,"Realized P&L %":round(np_/bv*100,2),
        "Running Capital":round(capital,2),
        "Trade Duration":f"{_dur_min(bt,st_)}m",
        "Market Regime":"REVERSAL","Trade Outcome":"WIN" if np_>0 else "LOSS"})
    tid+=1

print(f"Total trades: {len(trades)}")
df=pd.DataFrame(trades)
df["_dt"]=pd.to_datetime(df["Trade Date"],format="%d-%b-%Y")
df["_mo"]=df["_dt"].dt.to_period("M").dt.to_timestamp()


def build_monthly(df):
    rows=[]; ob=4_800_000.0
    for m in sorted(df["_mo"].unique()):
        mdf=df[df["_mo"]==m]
        cl=mdf["Running Capital"].iloc[-1]
        wins=mdf[mdf["Trade Outcome"]=="WIN"]
        loss=mdf[mdf["Trade Outcome"]=="LOSS"]
        cap_s=mdf["Running Capital"].values
        pk,mdd=cap_s[0],0.0
        for c in cap_s:
            if c>pk: pk=c
            if (pk-c)/pk*100>mdd: mdd=(pk-c)/pk*100
        rows.append({"Month":m.strftime("%b-%Y"),"Opening Balance":round(ob,2),
            "Closing Balance":round(cl,2),"Gross P&L":round(mdf["Gross P&L"].sum(),2),
            "Charges":round(mdf["Total Charges"].sum(),2),"Net P&L":round(mdf["Net P&L"].sum(),2),
            "Winning Trades":len(wins),"Losing Trades":len(loss),
            "Win Rate %":round(len(wins)/len(mdf)*100,1),
            "Average Winner":round(wins["Net P&L"].mean(),2) if len(wins)>0 else 0,
            "Average Loser":round(loss["Net P&L"].mean(),2) if len(loss)>0 else 0,
            "Largest Winner":round(wins["Net P&L"].max(),2) if len(wins)>0 else 0,
            "Largest Loser":round(loss["Net P&L"].min(),2) if len(loss)>0 else 0,
            "Max Drawdown %":round(mdd,2),
            "Return %":round((cl-ob)/ob*100,2)})
        ob=cl
    return pd.DataFrame(rows)


def build_perf(df):
    wins=df[df["Trade Outcome"]=="WIN"]; loss=df[df["Trade Outcome"]=="LOSS"]
    n=len(df); wr=round(len(wins)/n*100,2)
    pf=round(wins["Net P&L"].sum()/abs(loss["Net P&L"].sum()),2) if loss["Net P&L"].sum()!=0 else 99.0
    dr=df.groupby("_dt")["Net P&L"].sum()/4_800_000
    sh=round(dr.mean()/dr.std()*(252**0.5),2) if dr.std()>0 else 0.0
    cap=df["Running Capital"].values; pk,mdd_p,mdd_r=cap[0],0.0,0.0
    for c in cap:
        if c>pk: pk=c
        if (pk-c)/pk*100>mdd_p: mdd_p=(pk-c)/pk*100; mdd_r=pk-c
    outs=df["Trade Outcome"].tolist(); mw=ml=cw=cl=0
    for o in outs:
        if o=="WIN": cw+=1; cl=0
        else: cl+=1; cw=0
        mw=max(mw,cw); ml=max(ml,cl)
    aw=wins["Net P&L"].mean() if len(wins)>0 else 0
    al=abs(loss["Net P&L"].mean()) if len(loss)>0 else 1
    exp=round((wr/100*aw)-((1-wr/100)*al),2)
    rows=[("Total Trades",n),("Win Rate %",wr),("Profit Factor",pf),
          ("Sharpe Ratio (annualized)",sh),("Max Drawdown ₹",round(mdd_r,2)),
          ("Max Drawdown %",round(mdd_p,2)),("Average Trade ₹",round(df["Net P&L"].mean(),2)),
          ("Expectancy ₹",exp),("Largest Win ₹",round(wins["Net P&L"].max(),2)),
          ("Largest Loss ₹",round(loss["Net P&L"].min(),2)),
          ("Longest Win Streak",mw),("Longest Loss Streak",ml),
          ("Total Gross P&L ₹",round(df["Gross P&L"].sum(),2)),
          ("Total Charges ₹",round(df["Total Charges"].sum(),2)),
          ("Total Net P&L ₹",round(df["Net P&L"].sum(),2)),
          ("Starting Capital ₹",4_800_000),
          ("Ending Capital ₹",round(df["Running Capital"].iloc[-1],2)),
          ("Overall Return %",round((df["Running Capital"].iloc[-1]-4_800_000)/4_800_000*100,2))]
    perf=pd.DataFrame(rows,columns=["Metric","Value"])
    mr=df.groupby("_mo")["Net P&L"].sum().reset_index()
    mr.columns=["_mo","Net P&L"]; mr["Month"]=mr["_mo"].dt.strftime("%b-%Y")
    return perf,mr[["Month","Net P&L"]]


monthly_df=build_monthly(df); perf_df,mr_df=build_perf(df)
print("\nMonthly Summary:")
print(monthly_df[["Month","Opening Balance","Closing Balance","Win Rate %","Return %"]].to_string())
print("\nPerformance:")
print(perf_df.to_string())

# ─── EXCEL ──────────────────────────────────────────────────────────────────────
OUT_XL  = "/Users/sohelpathan/BuildSoft/Zerodha_web/Synthetic_Options_Trading_Dataset.xlsx"
OUT_CSV = "/Users/sohelpathan/BuildSoft/Zerodha_web/Synthetic_Options_Trades.csv"

HF=PatternFill("solid",fgColor="1F3864"); HN=Font(color="FFFFFF",bold=True,size=10)
WF=PatternFill("solid",fgColor="E2EFDA"); LF=PatternFill("solid",fgColor="FCE4D6")
AF=PatternFill("solid",fgColor="F2F2F2")
BD=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
MONEY={"Buy Value","Sell Value","Gross P&L","Brokerage","STT","Exchange Charges",
       "GST","SEBI Charges","Stamp Duty","Total Charges","Net P&L","Running Capital"}

def hdr(ws,row,n):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c)
        cell.font=HN; cell.fill=HF
        cell.alignment=Alignment(horizontal="center",wrap_text=True); cell.border=BD

wb=openpyxl.Workbook()
ws1=wb.active; ws1.title="Trades"
ws1.merge_cells("A1:AB1")
c=ws1["A1"]
c.value="⚠ SYNTHETIC DATASET FOR BACKTESTING AND EDUCATIONAL PURPOSES ONLY"
c.font=Font(bold=True,color="FF0000",size=11); c.alignment=Alignment(horizontal="center")
ws1.append([])
tcols=[col for col in df.columns if not col.startswith("_")]
ws1.append(tcols); hdr(ws1,ws1.max_row,len(tcols))
for row in df[tcols].itertuples(index=False):
    r=list(row); rfill=WF if r[tcols.index("Trade Outcome")]=="WIN" else LF
    ws1.append(r); rn=ws1.max_row
    for ci,(v,cn) in enumerate(zip(r,tcols),1):
        cell=ws1.cell(row=rn,column=ci); cell.fill=rfill; cell.border=BD
        cell.alignment=Alignment(horizontal="center")
        if cn in MONEY: cell.number_format="#,##0.00"
for i,w in enumerate([10,12,10,12,32,12,8,8,8,10,10,10,14,14,14,
                       12,10,14,10,12,12,14,14,12,16,12,14,10][:len(tcols)],1):
    ws1.column_dimensions[get_column_letter(i)].width=w
ws1.freeze_panes="A4"

ws2=wb.create_sheet("Monthly Summary")
ws2.merge_cells("A1:O1"); ws2["A1"].value="MONTHLY PERFORMANCE SUMMARY — SYNTHETIC DATASET"
ws2["A1"].font=Font(bold=True,size=12,color="1F3864"); ws2["A1"].alignment=Alignment(horizontal="center")
ws2.append([]); ws2.append(list(monthly_df.columns)); hdr(ws2,ws2.max_row,len(monthly_df.columns))
for row in monthly_df.itertuples(index=False):
    ws2.append(list(row)); rn=ws2.max_row; ret=list(row)[-1]
    rfill=WF if ret>=0 else LF
    for ci in range(1,len(monthly_df.columns)+1):
        cell=ws2.cell(row=rn,column=ci); cell.fill=rfill; cell.border=BD
        cell.alignment=Alignment(horizontal="center")
        if isinstance(cell.value,float): cell.number_format="#,##0.00"
for i,w in enumerate([12,18,18,16,14,14,12,12,10,14,14,16,16,12,10],1):
    ws2.column_dimensions[get_column_letter(i)].width=w
ws2.freeze_panes="A4"

ws3=wb.create_sheet("Performance Analytics")
ws3.merge_cells("A1:C1"); ws3["A1"].value="PERFORMANCE ANALYTICS — SYNTHETIC DATASET"
ws3["A1"].font=Font(bold=True,size=12,color="1F3864"); ws3["A1"].alignment=Alignment(horizontal="center")
ws3.append([]); ws3.append(["Metric","Value"]); hdr(ws3,ws3.max_row,2)
for ri,row in enumerate(perf_df.itertuples(index=False),ws3.max_row+1):
    ws3.append(list(row)); rn=ws3.max_row
    for ci in range(1,3):
        cell=ws3.cell(row=rn,column=ci); cell.border=BD
        if ri%2==0: cell.fill=AF
        if ci==1: cell.font=Font(bold=True)
        if isinstance(cell.value,float): cell.number_format="#,##0.00"
ws3.append([]); ws3.append(["MONTHLY RETURNS"]); ws3.cell(ws3.max_row,1).font=Font(bold=True,size=11,color="1F3864")
ws3.append(["Month","Net P&L ₹"]); hdr(ws3,ws3.max_row,2)
for row in mr_df.itertuples(index=False):
    ws3.append(list(row)); rn=ws3.max_row
    v=ws3.cell(row=rn,column=2).value; rfill=WF if isinstance(v,(int,float)) and v>=0 else LF
    for ci in range(1,3):
        cell=ws3.cell(row=rn,column=ci); cell.fill=rfill; cell.border=BD
        if isinstance(cell.value,float): cell.number_format="#,##0.00"
ws3.column_dimensions["A"].width=36; ws3.column_dimensions["B"].width=22
wb.save(OUT_XL); print(f"\n✅ Excel: {OUT_XL}")
df[tcols].to_csv(OUT_CSV,index=False); print(f"✅ CSV:   {OUT_CSV}")

# ─── VALIDATION ─────────────────────────────────────────────────────────────────
print("\n"+"="*65+"\nVALIDATION REPORT\n"+"="*65)
n=len(trades)
print(f"{'✓' if 500<=n<=700 else '⚠'} Total trades: {n}  (target 500–700)")
cap2=4_800_000.0
for t in trades: cap2+=t["Net P&L"]
diff=abs(cap2-trades[-1]["Running Capital"])
print(f"✓ Capital reconciles: ₹{diff:.2f}  {'PASS' if diff<50 else 'CHECK'}")
bad=sum(1 for t in trades if t["Quantity"]%LOT_SIZES[t["Instrument"]]!=0)
print(f"✓ Lot sizes: {'PASS' if bad==0 else f'FAIL({bad})'}")
bad2=sum(1 for t in trades if datetime.strptime(t["Trade Date"],"%d-%b-%Y").date()>datetime.strptime(t["Expiry Date"],"%d-%b-%Y").date())
print(f"✓ Expiry validity: {'PASS' if bad2==0 else f'FAIL({bad2})'}")
bad3=sum(1 for t in trades if abs(round(t["Brokerage"]+t["STT"]+t["Exchange Charges"]+t["GST"]+t["SEBI Charges"]+t["Stamp Duty"],2)-t["Total Charges"])>0.05)
print(f"✓ Charges reconcile: {'PASS' if bad3==0 else f'FAIL({bad3})'}")
bad4=sum(1 for t in trades if abs(round(t["Gross P&L"]-t["Total Charges"],2)-t["Net P&L"])>0.05)
print(f"✓ Net P&L=Gross−Charges: {'PASS' if bad4==0 else f'FAIL({bad4})'}")
print("\nCapital Milestones:")
ddf=df.copy(); ddf["_dt2"]=pd.to_datetime(ddf["Trade Date"],format="%d-%b-%Y")
for d,target in MILESTONES[1:]:
    if d>end: continue
    nb=ddf[ddf["_dt2"]<=pd.Timestamp(d)]
    if len(nb)>0:
        actual=nb["Running Capital"].iloc[-1]; pct=abs(actual-target)/target*100
        print(f"  {'✓' if pct<30 else '⚠'} {d.strftime('%d-%b-%Y')}: target ₹{target:>14,.0f} | actual ₹{actual:>14,.0f} | diff {pct:.1f}%")
print("\n✅ SYNTHETIC label: YES\n"+"="*65+"\nDONE")
